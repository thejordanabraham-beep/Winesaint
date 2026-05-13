import openpyxl
import re

SRC = '/Users/jordanabraham/Desktop/WINESAINT_MSTR_RVW_CLEANED.xlsx'
SHEET = 'France - Champagne'

# Rows to delete entirely (magnum duplicates with standard counterpart)
DELETE_ROWS = {20, 425, 434, 481, 1361}

# All broken word patterns (combined from France - Other passes)
BROKEN_WORDS = {
    'Sch lumber ger': 'Schlumberger',
    'Pf in gst berg': 'Pfingstberg',
    'Mu en ch berg': 'Muenchberg',
    'gra uwa cke': 'grauwacke',
    'So nnenglanz': 'Sonnenglanz',
    'So mmerberg': 'Sommerberg',
    'Vend an ges': 'Vendanges',
    'Card in aux': 'Cardinaux',
    'Winds buhl': 'Windsbuhl',
    'bo try tis': 'botrytis',
    'de stemmed': 'destemmed',
    'We in bach': 'Weinbach',
    'Bu ec her': 'Buecher',
    'Sc him berg': 'Schimberg',
    'La pierre': 'Lapierre',
    'Domain es': 'Domaines',
    'Caracol es': 'Caracoles',
    'In terdit': 'Interdit',
    'Pierre ts': 'Pierrets',
    'Trim bach': 'Trimbach',
    'Foil lard': 'Foillard',
    'So edlen': 'Soedlen',
    'Mam bourg': 'Mambourg',
    'Sylvan er': 'Sylvaner',
    'Tis sot': 'Tissot',
    'Box ler': 'Boxler',
    'so ils': 'soils',
    'zest in ess': 'zestiness',
    'Zest in ess': 'Zestiness',
    'tang in ess': 'tanginess',
    'chalk in ess': 'chalkiness',
    'fruit in ess': 'fruitiness',
    'pet rich or': 'petrichor',
    'Stein grub ler': 'Steingrubler',
    're montage': 'remontage',
    'in tertwined': 'intertwined',
    'Mirabel le': 'Mirabelle',
    'mirabel le': 'mirabelle',
    'Tar dives': 'Tardives',
    'over tones': 'overtones',
    'over tone': 'overtone',
    'under stated': 'understated',
    'in cluding': 'including',
    'in tensifies': 'intensifies',
    'in tensify': 'intensify',
    'in tegrated': 'integrated',
    'in tegrate': 'integrate',
    'to uches': 'touches',
    'to uchof': 'touch of',
    'to uching': 'touching',
    'GrandCru': 'Grand Cru',
    'not ions': 'notions',
    'of fering': 'offering',
    'of fered': 'offered',
    'of fers': 'offers',
    'of fera': 'offer a',
    'to taled': 'totaled',
    'to nes': 'tones',
    'to ned': 'toned',
    'mouth feel': 'mouthfeel',
}

def fix_text(text):
    if not text:
        return text

    for broken, fixed in BROKEN_WORDS.items():
        text = text.replace(broken, fixed)

    # Fix missing space after period before capital letter
    text = re.sub(r'(?<![A-Z])\.([A-Z])', r'. \1', text)

    # Fix word-number stuck (letter immediately before 3-4 digit number)
    text = re.sub(r'([a-z])(\d{3,4})(?=\s|[,.\-]|$)', r'\1 \2', text)

    # Fix "is100%" pattern specifically
    text = re.sub(r'is(\d+%)', r'is \1', text)

    # Fix "Read More" scraping artifacts
    text = re.sub(r'Read More\d{4}', '', text)
    text = re.sub(r'Readers\d{4}', '', text)

    # Clean up magnum format references in notes
    magnum_patterns = [
        r'\s*Available only in magnum format\.\s*',
        r'\s*Magnum-only release\.\s*',
        r'\s*The magnum format (?:brings|adds|delivers|provides)[^.]*\.\s*',
        r',?\s*tasted from magnum,?\s*',
        r',?\s*Tasted from magnum,?\s*',
    ]
    for pat in magnum_patterns:
        text = re.sub(pat, ' ', text)

    text = re.sub(r'\s{2,}', ' ', text).strip()
    return text

def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb[SHEET]

    # Collect rows to delete (process in reverse to keep indices stable)
    rows_to_delete = sorted(DELETE_ROWS, reverse=True)

    stats = {'deleted': 0, 'renamed': 0, 'notes_fixed': 0}

    # First pass: fix content
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value is None:
            continue

        name_cell = row[1]
        notes_cell = row[9]

        # Strip (Magnum)/(magnum) from wine names (for RENAME rows)
        if name_cell.value:
            old = str(name_cell.value)
            new = re.sub(r'\s*\((?:M|m)agnum\)', '', old).strip()
            if new != old:
                name_cell.value = new
                stats['renamed'] += 1
                print(f'  Renamed row {row_idx}: "{old}" → "{new}"')

        # Fix tasting notes
        if notes_cell.value:
            old = str(notes_cell.value)
            new = fix_text(old)
            if new != old:
                notes_cell.value = new
                stats['notes_fixed'] += 1

    # Delete duplicate magnum rows
    for row_idx in rows_to_delete:
        r = ws[row_idx]
        name = r[1].value
        print(f'  Deleting row {row_idx}: "{name}"')
        ws.delete_rows(row_idx, 1)
        stats['deleted'] += 1

    wb.save(SRC)
    print(f'\n=== CHAMPAGNE CLEANUP SUMMARY ===')
    print(f'Rows deleted (magnum dupes): {stats["deleted"]}')
    print(f'Wine names renamed (magnum stripped): {stats["renamed"]}')
    print(f'Tasting notes fixed: {stats["notes_fixed"]}')

if __name__ == '__main__':
    main()
