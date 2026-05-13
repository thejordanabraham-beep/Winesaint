import openpyxl
import re
from collections import defaultdict, Counter

SRC = '/Users/jordanabraham/Desktop/WINESAINT_MSTR_RVW_CLEANED.xlsx'

TABS = [
    'California', 'Austria', 'Australia', 'France - Bordeaux',
    'France - Burgundy', 'France - Loire', 'France - Rhone',
    'Germany', 'New Zealand', 'Oregon', 'South Africa', 'Spain', 'Italy',
]

# Known broken word patterns
KNOWN_BROKEN = [
    'in ess', 'over tones', 'over tone', 'under stated', 'de stemmed',
    'bo try tis', 'mouth feel', 'pet rich or', 're montage', 'not ions',
    'of fering', 'of fered', 'of fers', 'to uches', 'to nes', 'to ned',
    'so ils', 'so ars', 'in cluding', 'in tegrated', 'in tegrate',
    'in tensif', 'to uching', 'to uchof', 'of fera',
    'Mirabel le', 'mirabel le', 'Tar dives',
    'GrandCru', 'to taled',
]

# Common English 2-letter words to exclude from broken word scan
COMMON_2LETTER = {
    'an', 'as', 'at', 'be', 'by', 'do', 'go', 'he', 'if', 'in', 'is',
    'it', 'me', 'my', 'no', 'of', 'on', 'or', 'so', 'to', 'up', 'us',
    'we', 'am', 'ah', 'oh', 'ok', 'ha', 'hi',
}

# Common legitimate "2-letter + word" combos to exclude
LEGIT_COMBOS = {
    'an excellent', 'an elegant', 'an impressive', 'an attractive',
    'an underlying', 'an aromatic', 'an exceptional', 'an intense',
    'an old', 'an earthy', 'an expression', 'an abundance', 'an almost',
    'an apple', 'an absolutely', 'an amazing', 'an opulent',
    'is very', 'is quite', 'is really', 'is still', 'is rich',
    'is more', 'is well', 'is also', 'is not', 'is laced', 'is deep',
    'is full', 'is pure', 'is the', 'is one', 'is all',
    'at its', 'at the', 'at once', 'at least', 'at first',
    'by the', 'by its', 'by far', 'by ripe', 'by soft',
    'do not', 'do well',
    'go with', 'go well',
    'he says', 'he has',
    'if the', 'if not', 'if you', 'if anything',
    'in the', 'in its', 'in this', 'in fact', 'in oak', 'in new',
    'in old', 'in red', 'in all', 'in one', 'in ripe', 'in some',
    'in stainless', 'in steel', 'in bottle', 'in barrel', 'in wood',
    'in concrete', 'in french', 'in used', 'in large', 'in small',
    'in half', 'in whole', 'in part', 'in equal', 'in both',
    'it has', 'it shows', 'it will', 'it was', 'it needs',
    'me think',
    'my tasting',
    'no exception', 'no shortage', 'no doubt', 'no hurry', 'no rush',
    'of the', 'of its', 'of this', 'of ripe', 'of red', 'of dark',
    'of black', 'of rich', 'of oak', 'of new', 'of old', 'of fine',
    'of pure', 'of sweet', 'of fresh', 'of spice', 'of fruit',
    'of very', 'of wild', 'of warm', 'of wet', 'of dried', 'of soft',
    'of crisp', 'of deep', 'of well', 'of good', 'of great', 'of white',
    'of cherry', 'of citrus', 'of mineral', 'of floral',
    'on the', 'on its', 'on palate',
    'or so', 'or more', 'or two', 'or three',
    'so much', 'so well', 'so very', 'so young', 'so many', 'so rich',
    'so deep', 'so fine', 'so much', 'so lovely', 'so good', 'so pure',
    'to the', 'to its', 'to age', 'to be', 'to drink', 'to show',
    'to this', 'to open', 'to give', 'to have', 'to make', 'to match',
    'to try', 'to get', 'to find', 'to see', 'to say', 'to keep',
    'to pair', 'to come', 'to enjoy', 'to develop', 'to taste',
    'to ten', 'to twelve', 'to fifteen', 'to twenty',
    'up with', 'up the', 'up its', 'up front', 'up nicely',
    'us that', 'us with',
    'we have', 'we see', 'we get',
}


def check_tab(ws, sheet_name):
    stats = {
        'producer_prefix': 0,
        'broken_words': 0,
        'period_nospace': 0,
        'comma_nospace': 0,
        'word_number_stuck': 0,
        'read_more': 0,
        'score_discrepancy': 0,
        'name_junk': 0,
        'accent_inconsistencies': 0,
    }

    broken_counter = Counter()
    new_broken_counter = Counter()
    producer_prefix_samples = []
    score_disc_samples = []
    name_junk_samples = []

    # For accent inconsistency check
    name_groups = defaultdict(list)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value is None:
            continue

        producer = str(row[0].value).strip() if row[0].value else ''
        name = str(row[1].value).strip() if row[1].value else ''
        score = row[3].value if len(row) > 3 else None
        notes = str(row[9].value) if len(row) > 9 and row[9].value else ''

        # 1. Producer prefix in notes
        if notes:
            if re.match(r'^/', notes) or re.search(r'\d{2}/', notes[:20]):
                stats['producer_prefix'] += 1
                if len(producer_prefix_samples) < 3:
                    producer_prefix_samples.append((row_idx, notes[:80]))

        # 2. Broken words
        if notes:
            for pattern in KNOWN_BROKEN:
                count = notes.count(pattern)
                if count > 0:
                    stats['broken_words'] += count
                    broken_counter[pattern] += count

            # Scan for new potential broken words (2-letter + space + 3+ letters)
            for m in re.finditer(r'\b([a-z]{2}) ([a-z]{3,})\b', notes):
                combo = m.group(0)
                if combo.lower() not in LEGIT_COMBOS and m.group(1) not in COMMON_2LETTER:
                    new_broken_counter[combo] += 1

        # 3. Missing space after period
        if notes:
            hits = re.findall(r'(?<![A-Z])\.([A-Z])', notes)
            stats['period_nospace'] += len(hits)

        # 4. Missing space after comma
        if notes:
            hits = re.findall(r',([A-Za-z])', notes)
            stats['comma_nospace'] += len(hits)

        # 5. Word-number stuck
        if notes:
            hits = re.findall(r'[a-z]\d{3,4}(?=\s|[,.\-]|$)', notes)
            stats['word_number_stuck'] += len(hits)

        # 6. Read More artifacts
        if notes:
            hits = re.findall(r'Read More\d{4}|Readers\d{4}', notes)
            stats['read_more'] += len(hits)

        # 7. Score discrepancy
        if notes:
            m = re.match(r'^(\d{2,3})\b', notes.strip())
            if m:
                note_score = int(m.group(1))
                if score is not None:
                    try:
                        actual_score = int(float(score))
                        if note_score != actual_score and 80 <= note_score <= 100:
                            stats['score_discrepancy'] += 1
                            if len(score_disc_samples) < 3:
                                score_disc_samples.append((row_idx, actual_score, note_score, notes[:60]))
                    except (ValueError, TypeError):
                        pass

        # 8. Wine name junk
        if name:
            if len(name) > 100 or name.startswith('/') or '<' in name or 'http' in name.lower():
                stats['name_junk'] += 1
                if len(name_junk_samples) < 3:
                    name_junk_samples.append((row_idx, name[:80]))

        # 9. Accent inconsistency grouping
        if name:
            norm = name.lower()
            replacements = {
                'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
                'à': 'a', 'â': 'a', 'î': 'i', 'ï': 'i',
                'ô': 'o', 'ù': 'u', 'û': 'u', 'ü': 'u',
                'ç': 'c', 'ä': 'a', 'ö': 'o', 'ñ': 'n',
                'ß': 'ss',
            }
            for a, b in replacements.items():
                norm = norm.replace(a, b)
            norm = re.sub(r'[^\w\s-]', '', norm)
            norm = re.sub(r'\s+', ' ', norm).strip()
            name_groups[(producer, norm)].append((row_idx, name))

    # Count accent inconsistencies
    for (prod, norm), entries in name_groups.items():
        if len(entries) < 2:
            continue
        unique_names = set(e[1] for e in entries)
        if len(unique_names) > 1:
            stats['accent_inconsistencies'] += 1

    return stats, broken_counter, new_broken_counter, producer_prefix_samples, score_disc_samples, name_junk_samples


def main():
    wb = openpyxl.load_workbook(SRC)

    all_broken = Counter()
    all_new_broken = Counter()
    all_stats = {}
    all_samples = {}

    for tab in TABS:
        ws = wb[tab]
        stats, broken, new_broken, pp_samples, sd_samples, nj_samples = check_tab(ws, tab)
        all_stats[tab] = stats
        all_broken += broken
        all_new_broken += new_broken
        all_samples[tab] = {
            'producer_prefix': pp_samples,
            'score_discrepancy': sd_samples,
            'name_junk': nj_samples,
        }

    # Print summary table
    print("=" * 140)
    print(f"{'Tab':25s} | {'ProdPfx':>7s} | {'Broken':>7s} | {'Per.NoSp':>8s} | {'Com.NoSp':>8s} | {'WdNum':>5s} | {'RdMore':>6s} | {'ScDisc':>6s} | {'NmJunk':>6s} | {'Accent':>6s}")
    print("-" * 140)

    totals = defaultdict(int)
    for tab in TABS:
        s = all_stats[tab]
        print(f"{tab:25s} | {s['producer_prefix']:7d} | {s['broken_words']:7d} | {s['period_nospace']:8d} | {s['comma_nospace']:8d} | {s['word_number_stuck']:5d} | {s['read_more']:6d} | {s['score_discrepancy']:6d} | {s['name_junk']:6d} | {s['accent_inconsistencies']:6d}")
        for k, v in s.items():
            totals[k] += v

    print("-" * 140)
    print(f"{'TOTAL':25s} | {totals['producer_prefix']:7d} | {totals['broken_words']:7d} | {totals['period_nospace']:8d} | {totals['comma_nospace']:8d} | {totals['word_number_stuck']:5d} | {totals['read_more']:6d} | {totals['score_discrepancy']:6d} | {totals['name_junk']:6d} | {totals['accent_inconsistencies']:6d}")
    print("=" * 140)

    # Top 20 known broken word patterns
    print("\n=== TOP 20 KNOWN BROKEN WORD PATTERNS ===")
    for pattern, count in all_broken.most_common(20):
        print(f"  '{pattern}': {count}")

    # Top 30 potential NEW broken word patterns
    print("\n=== TOP 30 POTENTIAL NEW BROKEN PATTERNS (2-letter + word) ===")
    for pattern, count in all_new_broken.most_common(30):
        if count >= 3:
            print(f"  '{pattern}': {count}")

    # Samples
    print("\n=== PRODUCER PREFIX SAMPLES ===")
    for tab in TABS:
        samples = all_samples[tab]['producer_prefix']
        if samples:
            print(f"\n  [{tab}]")
            for row_idx, text in samples:
                print(f"    row {row_idx}: \"{text}...\"")

    print("\n=== SCORE DISCREPANCY SAMPLES ===")
    for tab in TABS:
        samples = all_samples[tab]['score_discrepancy']
        if samples:
            print(f"\n  [{tab}]")
            for row_idx, actual, note_score, text in samples:
                print(f"    row {row_idx}: col_score={actual} note_score={note_score} \"{text}...\"")

    print("\n=== NAME JUNK SAMPLES ===")
    for tab in TABS:
        samples = all_samples[tab]['name_junk']
        if samples:
            print(f"\n  [{tab}]")
            for row_idx, text in samples:
                print(f"    row {row_idx}: \"{text}\"")

    # Per-tab accent inconsistency details
    print("\n=== ACCENT INCONSISTENCY DETAILS (top 5 per tab) ===")
    for tab in TABS:
        ws = wb[tab]
        name_groups = defaultdict(list)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value is None or row[1].value is None:
                continue
            producer = str(row[0].value).strip()
            name = str(row[1].value).strip()
            norm = name.lower()
            replacements = {
                'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
                'à': 'a', 'â': 'a', 'î': 'i', 'ï': 'i',
                'ô': 'o', 'ù': 'u', 'û': 'u', 'ü': 'u',
                'ç': 'c', 'ä': 'a', 'ö': 'o', 'ñ': 'n', 'ß': 'ss',
            }
            for a, b in replacements.items():
                norm = norm.replace(a, b)
            norm = re.sub(r'[^\w\s-]', '', norm)
            norm = re.sub(r'\s+', ' ', norm).strip()
            name_groups[(producer, norm)].append((row_idx, name))

        inconsistent = []
        for (prod, norm), entries in name_groups.items():
            if len(entries) < 2:
                continue
            unique_names = set(e[1] for e in entries)
            if len(unique_names) > 1:
                inconsistent.append((prod, unique_names, len(entries)))

        if inconsistent:
            print(f"\n  [{tab}] — {len(inconsistent)} groups")
            for prod, names, count in inconsistent[:5]:
                print(f"    [{prod}] ({count} wines): {sorted(names)}")


if __name__ == '__main__':
    main()
